from schemas.query.sort.report import ReportSort, SortDirection, SortReportBy
from schemas.shared.paginated_response import PaginatedResponse
from io import BytesIO
from openpyxl import load_workbook
import pytest

class TestReportServiceRead:
    def test_get_report_paginated_success(self, report_service, mock_user, mock_report):
        # Arrange
        sort = ReportSort(
            page=1,
            size=10,
            sort_by=SortReportBy.CATEGORY,
            sort_direction=SortDirection.ASC
        )
        mock_reports = [mock_report, mock_report]
        total = len(mock_reports)
        paginated_response = PaginatedResponse(
            data=mock_reports,
            meta={
                "page": sort.page,
                "page_size": sort.size,
                "total": total,
                "total_pages": (total + sort.size - 1) // sort.size
            }
        )
        report_service.repository.get_report_paginated.return_value = paginated_response

        # Act
        result = report_service.get_report_paginated(sort, mock_user)

        # Assert
        assert len(result.data) == 2
        assert result.meta.total == 2
        assert result.meta.page == 1
        report_service.repository.get_report_paginated.assert_called_once_with(
            sort, mock_user.location)

    def test_get_report_paginated_empty(self, report_service, mock_user):
        # Arrange
        sort = ReportSort(
            page=1,
            size=10,
            sort_by=SortReportBy.CATEGORY,
            sort_direction=SortDirection.ASC
        )
        paginated_response = PaginatedResponse(
            data=[],
            meta={
                "page": sort.page,
                "page_size": sort.size,
                "total": 0,
                "total_pages": 0
            }
        )
        report_service.repository.get_report_paginated.return_value = paginated_response

        # Act
        result = report_service.get_report_paginated(sort, mock_user)

        # Assert
        assert len(result.data) == 0
        assert result.meta.total == 0
        assert result.meta.page == 1
        report_service.repository.get_report_paginated.assert_called_once_with(
            sort, mock_user.location)

    @pytest.mark.asyncio
    async def test_convert_to_excel(self, report_service, mock_user, mock_report):
        mock_report_1 = mock_report.copy(deep=True)
        mock_report_1.category = "Category 1"
        mock_report_2 = mock_report.copy(deep=True)
        mock_report_2.category = "Category 2"
        paginated_response_1 = PaginatedResponse(
            data=[
                mock_report_1,
                mock_report_2,
            ],
            meta={
                "page": 1,
                "page_size": 2,
                "total": 2,
                "total_pages": 1
            }
        )
        paginated_response_2 = PaginatedResponse(
            data=[],
            meta={
                "page": 2,
                "page_size": 2,
                "total": 2,
                "total_pages": 1
            }
        )
        report_service.repository.get_report_paginated.side_effect = [
            paginated_response_1,
            paginated_response_2
        ]

        excel_stream = await report_service.convert_to_excel(mock_user)
        assert isinstance(excel_stream, BytesIO)

        excel_stream.seek(0)
        workbook = load_workbook(excel_stream, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0] == ("Category", "Total", "Assigned", "Available",
                       "Not available", "Waiting for recycling", "Recycled")
        
        assert rows[1] == (
            mock_report_1.category,
            mock_report_1.total,
            mock_report_1.assigned,
            mock_report_1.available,
            mock_report_1.not_available,
            mock_report_1.waiting_for_recycling,
            mock_report_1.recycled
        )

        assert rows[2] == (
            mock_report_2.category,
            mock_report_2.total,
            mock_report_2.assigned,
            mock_report_2.available,
            mock_report_2.not_available,
            mock_report_2.waiting_for_recycling,
            mock_report_2.recycled
        )